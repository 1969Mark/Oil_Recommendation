"""
process_nb.py
=============
增量處理 NB_data/ 內的 PDF 檔案，輸出 output/nb_master.xlsx 及 parse_report。
依 CLAUDE.md 規格：Format A/B/C 自動偵測、信心分數、增量 registry。
"""

import os, sys
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _config import (
    NB_DIR, OUT_DIR, RPT_DIR,
    HDR_BG, NB_COLOR, COLS, DEDUP_KEYS,
    NB_MASTER as MASTER, NB_REG as REG_FILE,
)
from _common import (
    load_registry, save_registry, load_failed_registry,
    save_failed_registry, needs_processing, file_metadata,
    append_log, write_styled_sheet,
)
from _filters import (
    is_invalid_model, canonicalize_column, maker_key, model_key,
    strip_quantity_descriptor, strip_power_spec_parens,
    apply_part_semantic_merge, apply_compressor_part_rule, strip_non_fuel_parens,
)
from _nb_parsers import parse_pdf, assign_confidence, get_ship_id
from _dict_validator import build_known_sets, annotate_records, summarize_hits

# ── Excel 輸出 ───────────────────────────────────────────────────
def write_master_excel(df, path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'nb_master'
    out_cols = COLS + ['Source', 'source_file']
    write_styled_sheet(ws, df, out_cols, source_color=NB_COLOR)
    wb.save(path)

def write_parse_report(records_with_conf, fname):
    """產生 parse_report Excel"""
    rpt_path = os.path.join(RPT_DIR, f'nb_{os.path.splitext(fname)[0]}_report.xlsx')
    wb = Workbook()

    # Sheet 1: all_data（含字典命中標記欄）
    ws1 = wb.active
    ws1.title = 'all_data'
    report_cols = ['row_id', 'page_no', 'confidence', 'confidence_reason',
                   'Equipment', 'Maker', 'Model / Type', 'Part to be lubricated', 'Lubricant',
                   '_hit_Equipment', '_hit_Maker', '_hit_Part to be lubricated', '_hit_Lubricant']
    for ci, col in enumerate(report_cols, 1):
        c = ws1.cell(row=1, column=ci, value=col)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=HDR_BG)

    for ri, r in enumerate(records_with_conf, start=1):
        vals = [ri, r.get('_page', ''), r.get('confidence', ''), r.get('confidence_reason', ''),
                r.get('設備名稱', ''), r.get('設備廠家', ''), r.get('設備型號', ''),
                r.get('潤滑部位', ''), r.get('推薦潤滑油', ''),
                r.get('_hit_Equipment', ''), r.get('_hit_Maker', ''),
                r.get('_hit_Part to be lubricated', ''), r.get('_hit_Lubricant', '')]
        for ci, v in enumerate(vals, 1):
            ws1.cell(row=ri + 1, column=ci, value=str(v) if v else '')

    # Sheet 2: review_required
    ws2 = wb.create_sheet('review_required')
    review_cols = report_cols + ['review_status']
    for ci, col in enumerate(review_cols, 1):
        c = ws2.cell(row=1, column=ci, value=col)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='C00000')

    ri = 2
    for idx, r in enumerate(records_with_conf, start=1):
        if r.get('confidence') in ('medium', 'low'):
            vals = [idx, r.get('_page', ''), r.get('confidence', ''), r.get('confidence_reason', ''),
                    r.get('設備名稱', ''), r.get('設備廠家', ''), r.get('設備型號', ''),
                    r.get('潤滑部位', ''), r.get('推薦潤滑油', ''), '']
            for ci, v in enumerate(vals, 1):
                ws2.cell(row=ri, column=ci, value=str(v) if v else '')
            ri += 1

    wb.save(rpt_path)
    return rpt_path


# ── 主流程 ──────────────────────────────────────────────────────
def main():
    print("=" * 50)
    print("▶  process_nb.py 開始執行")
    print("=" * 50)

    reg  = load_registry(REG_FILE)
    freg = load_failed_registry()
    now  = datetime.now()

    all_pdfs = sorted([f for f in os.listdir(NB_DIR) if f.lower().endswith('.pdf')])
    if not all_pdfs:
        print(f"⚠️  找不到 PDF：{NB_DIR}")
        return

    to_process = []
    skipped    = []
    for fname in all_pdfs:
        fpath = os.path.join(NB_DIR, fname)
        needed, reason = needs_processing(fname, fpath, reg)
        if needed:
            to_process.append((fname, fpath, reason))
        else:
            skipped.append(fname)

    retryable = [
        k for k, v in freg.get('failed_files', {}).items()
        if v.get('source') == 'nb' and v.get('status') == 'retryable' and v.get('retry_count', 0) < 3
    ]
    for key in retryable:
        fname = os.path.basename(key)
        fpath = os.path.join(NB_DIR, fname)
        if os.path.exists(fpath) and fname not in [x[0] for x in to_process]:
            to_process.append((fname, fpath, 'retry'))

    print(f"PDF 總數：{len(all_pdfs)}  ｜  待處理：{len(to_process)}  ｜  跳過：{len(skipped)}")

    if not to_process:
        print("✓ 無需更新，所有檔案未變更。")
        return

    # 載入現有 master
    if os.path.exists(MASTER):
        try:
            df_master = pd.read_excel(MASTER, sheet_name='nb_master')
        except Exception:
            df_master = pd.DataFrame()
    else:
        df_master = pd.DataFrame()

    master_cols = COLS + ['Count', 'Source', 'source_file']

    stats = {'added': [], 'updated': [], 'failed': []}
    report_summary = []

    # 預載字典（一次性，給所有待處理檔共用）
    known_sets = build_known_sets()
    print(f"  字典規模：Equipment={len(known_sets.get('Equipment', set()))} "
          f"Maker={len(known_sets.get('Maker', set()))} "
          f"Part={len(known_sets.get('Part to be lubricated', set()))} "
          f"Lubricant={len(known_sets.get('Lubricant', set()))}")

    for fname, fpath, reason in to_process:
        try:
            print(f"\n  [{fname}]  [{reason}]")
            ship_id = get_ship_id(fname)
            records, fmt = parse_pdf(fpath, ship_id)
            print(f"    解析：{len(records)} 筆原始記錄")

            if not records:
                raise ValueError("EMPTY_DATA：解析後無任何有效資料列")

            # 信心分數
            records_with_conf = assign_confidence(records, fmt, fpath)

            # 字典命中標記（以現有 master 為字典反推欄位是否站對位置）
            records_with_conf = annotate_records(records_with_conf, known_sets)
            hit_sum = summarize_hits(records_with_conf)
            hit_str = '  '.join(f'{k}:{h}/{t}' for k, (h, t) in hit_sum.items())
            print(f'    字典命中  {hit_str}')

            # 產生 parse_report
            os.makedirs(RPT_DIR, exist_ok=True)
            rpt_path = write_parse_report(records_with_conf, fname)

            # 統計信心分數
            conf_count = {'high': 0, 'medium': 0, 'low': 0}
            for r in records_with_conf:
                conf_count[r.get('confidence', 'low')] = conf_count.get(r.get('confidence', 'low'), 0) + 1
            report_summary.append((fname, len(records_with_conf), conf_count))
            print(f"    信心分數  HIGH:{conf_count['high']}  MEDIUM:{conf_count['medium']}  LOW:{conf_count['low']}")

            # 轉換為 DataFrame（排除 low 信心分數）
            records_for_master = [r for r in records_with_conf if r.get('confidence') != 'low']
            low_count = len(records_with_conf) - len(records_for_master)
            if low_count > 0:
                print(f"    排除 low 信心分數：{low_count} 筆不納入 master")

            df_new = pd.DataFrame([{
                'Equipment': r.get('設備名稱', ''),
                'Maker': r.get('設備廠家', ''),
                'Model / Type': r.get('設備型號', ''),
                'Part to be lubricated': r.get('潤滑部位', ''),
                'Lubricant': r.get('推薦潤滑油', ''),
                'Source': 'NB',
                'source_file': fname,
            } for r in records_for_master])

            if df_new.empty:
                print(f"    ⚠️  過濾後無有效記錄，跳過")
                reg['processed_files'][fname] = {
                    'processed_at': now.isoformat(timespec='seconds'),
                    **file_metadata(fpath),
                }
                if reason == 'new':
                    stats['added'].append(f'{fname} ✓（無 high/medium 記錄）')
                else:
                    stats['updated'].append(f'{fname} ✓（無 high/medium 記錄）')
                continue

            # 全大寫
            for col in COLS:
                df_new[col] = df_new[col].fillna('').astype(str).str.strip().str.upper()

            # 從 master 移除舊資料
            if not df_master.empty and 'source_file' in df_master.columns:
                df_master = df_master[df_master['source_file'] != fname]

            df_master = pd.concat([df_master, df_new], ignore_index=True)

            # 更新 registry
            reg['processed_files'][fname] = {
                'processed_at': now.isoformat(timespec='seconds'),
                **file_metadata(fpath),
            }
            fail_key = f'nb/{fname}'
            if fail_key in freg.get('failed_files', {}):
                del freg['failed_files'][fail_key]

            if reason == 'retry':
                stats['added'].append(f'{fname} ✓（retry 成功）')
            elif reason == 'new':
                stats['added'].append(f'{fname} ✓')
            else:
                stats['updated'].append(f'{fname} ✓（{reason}）')

        except Exception as e:
            error_msg = str(e)
            print(f"  ❌ 失敗：{error_msg}")
            fail_key = f'nb/{fname}'
            error_type = 'EMPTY_DATA' if 'EMPTY_DATA' in error_msg else 'PDF_PARSE_ERROR'
            if fail_key not in freg.setdefault('failed_files', {}):
                freg['failed_files'][fail_key] = {
                    'source': 'nb', 'first_failed_at': now.isoformat(timespec='seconds'),
                    'last_failed_at': now.isoformat(timespec='seconds'),
                    'retry_count': 1, 'error_type': error_type,
                    'error_message': error_msg[:200], 'status': 'retryable'
                }
            else:
                freg['failed_files'][fail_key]['retry_count'] += 1
                freg['failed_files'][fail_key]['last_failed_at'] = now.isoformat(timespec='seconds')
                if freg['failed_files'][fail_key]['retry_count'] >= 3:
                    freg['failed_files'][fail_key]['status'] = 'manual_required'
            stats['failed'].append(fname)
            continue

    if df_master.empty:
        print("⚠️  無有效資料，跳過寫入。")
        return

    # 確保欄位完整
    for col in master_cols:
        if col not in df_master.columns:
            df_master[col] = ''

    # 移除 Model 末端數量描述
    before_m = df_master['Model / Type'].copy()
    df_master['Model / Type'] = df_master['Model / Type'].map(strip_quantity_descriptor)
    print(f"\n  Model 數量字尾移除：{(before_m != df_master['Model / Type']).sum()} 列")

    # 移除 Model 功率/轉速規格括號（含 KW/RPM/HP/PS/KVA/@）
    before_pow = df_master['Model / Type'].copy()
    df_master['Model / Type'] = df_master['Model / Type'].map(strip_power_spec_parens)
    print(f"  Model 功率規格括號移除：{(before_pow != df_master['Model / Type']).sum()} 列")

    # 過濾無效型號
    before = len(df_master)
    df_master = df_master[~df_master['Model / Type'].apply(is_invalid_model)]
    print(f"  過濾無效型號：{before - len(df_master)} 列移除")

    # 過濾空白 Lubricant
    before = len(df_master)
    df_master = df_master[df_master['Lubricant'].fillna('').astype(str).str.strip() != '']
    print(f"  過濾空白 Lubricant：{before - len(df_master)} 列移除")

    # 排除 TALUSIA LS 25
    before = len(df_master)
    df_master = df_master[~df_master['Lubricant'].str.contains('TALUSIA LS 25', na=False)]
    print(f"  排除 TALUSIA LS 25：{before - len(df_master)} 列移除")

    # 排除 NOT LUBRICATED
    before = len(df_master)
    df_master = df_master[df_master['Lubricant'].fillna('').astype(str).str.strip().str.upper() != 'NOT LUBRICATED']
    print(f"  排除 NOT LUBRICATED：{before - len(df_master)} 列移除")

    # Part 括號清理（保留燃油規範/EAL；移除註記、油路說明、方括號、ALTERNATE 後綴等）
    before_paren = df_master['Part to be lubricated'].copy()
    df_master['Part to be lubricated'] = df_master['Part to be lubricated'].map(strip_non_fuel_parens)
    print(f"  Part 括號清理：{(before_paren != df_master['Part to be lubricated']).sum()} 列改寫")

    # Equipment 含 COMPRESSOR → Part 統一為 CYLINDERS & BEARINGS
    before_cmp = df_master['Part to be lubricated'].copy()
    df_master['Part to be lubricated'] = df_master.apply(
        lambda r: apply_compressor_part_rule(r['Equipment'], r['Part to be lubricated']), axis=1)
    print(f"  COMPRESSOR Part 統一：{(before_cmp != df_master['Part to be lubricated']).sum()} 列改寫為 CYLINDERS & BEARINGS")

    # Part 語意合併（HYDRAULIC 同義詞 → HYDRAULIC SYSTEM；通用齒輪 → ENCLOSED GEAR）
    before_part = df_master['Part to be lubricated'].copy()
    df_master['Part to be lubricated'] = df_master['Part to be lubricated'].map(apply_part_semantic_merge)
    print(f"  Part 語意合併（HYDRAULIC/GEAR）：{(before_part != df_master['Part to be lubricated']).sum()} 列改寫")

    # Maker / Model / Part 正規化
    new_makers, mk_groups = canonicalize_column(df_master['Maker'], maker_key)
    new_models, md_groups = canonicalize_column(df_master['Model / Type'], model_key)
    new_parts,  pt_groups = canonicalize_column(df_master['Part to be lubricated'], model_key)
    df_master['Maker'] = new_makers
    df_master['Model / Type'] = new_models
    df_master['Part to be lubricated'] = new_parts
    print(f"  Maker 正規化：合併 {len(mk_groups)} 群組；Model 正規化：合併 {len(md_groups)} 群組；Part 正規化：合併 {len(pt_groups)} 群組")

    # 計算 Count
    df_master['Count'] = df_master.groupby(DEDUP_KEYS, dropna=False)['Maker'].transform('size').fillna(1).astype(int)

    # 去重
    before = len(df_master)
    df_master = df_master.drop_duplicates(subset=DEDUP_KEYS)
    print(f"  去重：{before - len(df_master)} 列移除，剩餘 {len(df_master)} 列")

    # 寫入 master
    try:
        os.makedirs(OUT_DIR, exist_ok=True)
        write_master_excel(df_master[master_cols], MASTER)
        print(f"\n✓ nb_master.xlsx 寫入完成：{len(df_master)} 列")
    except Exception as e:
        print(f"\n❌ WRITE_ERROR：{e}")
        return

    save_registry(reg, REG_FILE)
    save_failed_registry(freg)

    manual_req = [k for k, v in freg.get('failed_files', {}).items()
                  if v.get('source') == 'nb' and v.get('status') == 'manual_required']

    # Log
    log_lines = [
        '=' * 40,
        f'時間：{now.strftime("%Y-%m-%d %H:%M:%S")}',
        '動作：更新NB',
        '-' * 40,
    ]
    if stats['added']:   log_lines.append(f'新增處理：{", ".join(stats["added"])}')
    if stats['updated']: log_lines.append(f'重新處理：{", ".join(stats["updated"])}')
    if skipped:          log_lines.append(f'跳過：{len(skipped)} 個檔案（未變更）')
    if stats['failed']:
        log_lines.append('失敗（本次）：')
        for f in stats['failed']: log_lines.append(f'  - {f}')
    if manual_req:
        log_lines.append('\n⚠️  需人工介入（manual_required）：')
        for k in manual_req: log_lines.append(f'  - {k}')
    log_lines.append(f'\n輸出：output/nb_master.xlsx（共 {len(df_master)} 列）')
    log_lines.append('=' * 40)
    append_log('\n'.join(log_lines))

    # 摘要
    print('\n' + '=' * 50)
    print('📊 執行摘要')
    print('-' * 50)
    if stats['added']:   print(f'新增：{len(stats["added"])} 個檔案')
    if stats['updated']: print(f'更新：{len(stats["updated"])} 個檔案')
    if skipped:          print(f'跳過：{len(skipped)} 個檔案')
    if stats['failed']:  print(f'失敗：{len(stats["failed"])} 個檔案')
    print(f'輸出：nb_master.xlsx（{len(df_master):,} 列）')

    print('\n📋 解析報告已產生：')
    for fname, total, cc in report_summary:
        print(f'  - {fname} → 共 {total} 列')
        print(f'    HIGH: {cc["high"]} 列（無需核對）')
        if cc['medium'] > 0:
            rpt_name = f'nb_{os.path.splitext(fname)[0]}_report.xlsx'
            print(f'    MEDIUM: {cc["medium"]} 列（建議抄查）→ parse_report/{rpt_name}')
        if cc['low'] > 0:
            rpt_name = f'nb_{os.path.splitext(fname)[0]}_report.xlsx'
            print(f'    LOW: {cc["low"]} 列（需人工核對）→ parse_report/{rpt_name}')

    if manual_req:
        print('\n⚠️  以下檔案需要人工介入（已失敗 3 次）：')
        for k in manual_req: print(f'   - {k}')
    print('=' * 50)

if __name__ == '__main__':
    main()

