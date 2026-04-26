"""
_common.py
==========
三支 process 腳本的共用函式：檔案雜湊、Registry 讀寫、增量判斷、log 寫入、
Excel 樣式輸出。

常數（路徑、色彩、欄位）已集中至 _config.py；本模組亦從該處 re-export，
讓既有 `from _common import HDR_BG, ...` 寫法持續可用，避免大規模改動。
"""

import os
import json
import hashlib
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Re-export 常數（向後相容；新程式碼建議直接 from _config import ...）
from _config import (  # noqa: F401
    PROJECT, LC_DIR, NB_DIR, OEM_DIR, OUT_DIR, RPT_DIR, REG_DIR, LOG_DIR,
    LC_REG, NB_REG, OEM_REG, FAIL_REG, LOG_FILE,
    LC_MASTER, NB_MASTER, OEM_MASTER,
    PENDING_DIR, MANUAL_SRC, APP_HTML, APP_DATA_JS,
    HDR_BG, ODD_BG, EVEN_BG, LC_COLOR, NB_COLOR, OEM_COLOR, SOURCE_COLOR,
    COLS, DEDUP_KEYS, OEM_COLS, DISPLAY_COLS, SRC_ORDER,
)


# ── 檔案雜湊 ────────────────────────────────────────────────────
def sha256(path):
    """以 64 KB 區塊讀檔計算 sha256 摘要。"""
    h = hashlib.sha256()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(65536), b''):
            h.update(chunk)
    return h.hexdigest()


# ── Registry 讀寫 ───────────────────────────────────────────────
def load_registry(path):
    """讀取成功 registry；不存在時回傳空殼。"""
    if os.path.exists(path):
        with open(path, encoding='utf-8') as f:
            return json.load(f)
    return {'last_updated': '', 'processed_files': {}}


def save_registry(reg, path):
    """寫入成功 registry，並更新 last_updated 為現在時間。"""
    reg['last_updated'] = datetime.now().isoformat(timespec='seconds')
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(reg, f, indent=2, ensure_ascii=False)


def load_failed_registry():
    """讀取共用失敗 registry。"""
    if os.path.exists(FAIL_REG):
        with open(FAIL_REG, encoding='utf-8') as f:
            return json.load(f)
    return {'last_updated': '', 'failed_files': {}}


def save_failed_registry(freg):
    """寫入共用失敗 registry。"""
    freg['last_updated'] = datetime.now().isoformat(timespec='seconds')
    with open(FAIL_REG, 'w', encoding='utf-8') as f:
        json.dump(freg, f, indent=2, ensure_ascii=False)


def needs_processing(fname, fpath, reg):
    """判斷檔案是否需要重新處理：新檔 / mtime 異動 / sha256 異動。

    回傳：(need_process: bool, reason: str)
    reason ∈ {'new', 'mtime_changed', 'sha256_changed', 'unchanged'}
    """
    if fname not in reg['processed_files']:
        return True, 'new'
    rec = reg['processed_files'][fname]
    stat = os.stat(fpath)
    mtime = datetime.fromtimestamp(stat.st_mtime).isoformat(timespec='seconds')
    if mtime != rec.get('file_mtime', ''):
        return True, 'mtime_changed'
    if sha256(fpath) != rec.get('sha256', ''):
        return True, 'sha256_changed'
    return False, 'unchanged'


def file_metadata(fpath):
    """產生 registry record 用的 metadata：file_size_bytes / file_mtime / sha256。"""
    stat = os.stat(fpath)
    return {
        'file_size_bytes': stat.st_size,
        'file_mtime': datetime.fromtimestamp(stat.st_mtime).isoformat(timespec='seconds'),
        'sha256': sha256(fpath),
    }


# ── Log 寫入 ────────────────────────────────────────────────────
def append_log(summary):
    """Append 一段已組裝好的 log 文字到 logs/update_log.txt。"""
    os.makedirs(LOG_DIR, exist_ok=True)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(summary + '\n')


# ── Excel 樣式輸出 ──────────────────────────────────────────────
def write_styled_sheet(ws, df, columns, source_color=None, cell_styler=None):
    """以專案標準樣式寫入單一 worksheet。

    - 標題列：HDR_BG 背景 + 白色粗體 Arial 10
    - 資料列：奇數白 / 偶數淡藍 交錯，Arial 9
    - Source 欄：以 source_color 上色 + 粗體
    - 凍結窗格：A2（凍結標題列）
    - 欄寬：自動調整（10 ~ 50 字元）
    - cell_styler：可選 callback `(col_name, value, cell)`，
                  對特定欄位（例如 OEM 的 confidence）做進一步上色

    df: 來源 DataFrame
    columns: 要輸出的欄位順序（不在 df 中的欄位會留空）
    """
    ws.freeze_panes = 'A2'

    for ci, col in enumerate(columns, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        c.fill = PatternFill('solid', fgColor=HDR_BG)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for ri, (_, row) in enumerate(df.iterrows(), start=2):
        bg = ODD_BG if ri % 2 == 1 else EVEN_BG
        for ci, col in enumerate(columns, 1):
            v = row.get(col, '') if hasattr(row, 'get') else ''
            val = str(v) if pd.notna(v) else ''
            c = ws.cell(row=ri, column=ci, value=val)
            if col == 'Source' and source_color:
                c.font = Font(name='Arial', size=9, color=source_color, bold=True)
            else:
                c.font = Font(name='Arial', size=9)
            c.fill = PatternFill('solid', fgColor=bg)
            c.alignment = Alignment(vertical='center', wrap_text=True)
            if cell_styler:
                cell_styler(col, val, c)

    for ci, col in enumerate(columns, 1):
        if col in df.columns:
            max_len = max((len(str(v)) for v in df[col].fillna('') if v), default=len(col))
        else:
            max_len = len(col)
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len + 2, 10), 50)
