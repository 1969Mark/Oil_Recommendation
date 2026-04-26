"""
process_lube_chart.py
=====================
增量處理 LubeChart_data/ 內的 CSV 檔案，輸出 output/lube_chart_master.xlsx。
依 CLAUDE.md 規格：registry 比對、標準化、過濾、格式化 Excel 輸出。
"""

import os, sys, json, hashlib, glob, re
from datetime import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── 路徑設定 ────────────────────────────────────────────────────
PROJECT  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LC_DIR   = os.path.join(PROJECT, 'LubeChart_data')
OUT_DIR  = os.path.join(PROJECT, 'output')
REG_DIR  = os.path.join(PROJECT, 'registry')
LOG_DIR  = os.path.join(PROJECT, 'logs')
MASTER   = os.path.join(OUT_DIR, 'lube_chart_master.xlsx')
REG_FILE = os.path.join(REG_DIR, 'lube_chart_registry.json')
FAIL_REG = os.path.join(REG_DIR, 'failed_registry.json')
LOG_FILE = os.path.join(LOG_DIR, 'update_log.txt')

COLS = ['Equipment', 'Maker', 'Model / Type', 'Part to be lubricated', 'Lubricant']
DEDUP_KEYS = ['Maker', 'Model / Type', 'Part to be lubricated', 'Lubricant']
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _filters import is_invalid_model, canonicalize_column, maker_key, model_key, strip_quantity_descriptor

# ── Excel 色彩 ──────────────────────────────────────────────────
HDR_BG   = '1F3864'
ODD_BG   = 'FFFFFF'
EVEN_BG  = 'EBF3FB'
LC_COLOR = '375623'

# ── Registry 工具 ───────────────────────────────────────────────
def sha256(path):
    h = hashlib.sha256()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(65536), b''):
            h.update(chunk)
    return h.hexdigest()

def load_registry(path):
    if os.path.exists(path):
        with open(path, encoding='utf-8') as f:
            return json.load(f)
    return {'last_updated': '', 'processed_files': {}}

def save_registry(reg, path):
    reg['last_updated'] = datetime.now().isoformat(timespec='seconds')
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(reg, f, indent=2, ensure_ascii=False)

def load_failed_registry():
    if os.path.exists(FAIL_REG):
        with open(FAIL_REG, encoding='utf-8') as f:
            return json.load(f)
    return {'last_updated': '', 'failed_files': {}}

def save_failed_registry(freg):
    freg['last_updated'] = datetime.now().isoformat(timespec='seconds')
    with open(FAIL_REG, 'w', encoding='utf-8') as f:
        json.dump(freg, f, indent=2, ensure_ascii=False)

def needs_processing(fname, fpath, reg):
    """判斷檔案是否需要重新處理"""
    if fname not in reg['processed_files']:
        return True, 'new'
    rec = reg['processed_files'][fname]
    stat = os.stat(fpath)
    mtime = datetime.fromtimestamp(stat.st_mtime).isoformat(timespec='seconds')
    if mtime != rec.get('file_mtime', ''):
        return True, 'mtime_changed'
    if sha256(fpath) != rec.get('sha256', ''):
        return True, 'sha256_changed'
    return False, 'unchanged'

# ── Maker 標準化 ────────────────────────────────────────────────
def build_maker_norm_map(maker_series):
    freq = maker_series[maker_series != ''].value_counts()
    norm_map = {}
    for maker in freq.index:
        clean = re.sub(r'\s*\(?\s*[Xx]\s*\d+\s*\)?$', '', maker).strip()
        if clean != maker and clean and clean in freq and freq[clean] > freq[maker]:
            norm_map[maker] = clean
    for maker in freq.index:
        if maker in norm_map or '/' in maker or freq[maker] > 200:
            continue
        for base in freq.index:
            if base == maker or '/' in base:
                continue
            if (maker.startswith(base + ' ') or
                maker.startswith(base + '-') or
                maker.startswith(base + '(')):
                if freq[base] >= freq[maker] * 10:
                    norm_map[maker] = base
                    break
    return norm_map

# ── Part 標準化 ─────────────────────────────────────────────────
def build_part_norm_map(part_series):
    freq = part_series[part_series != ''].value_counts()
    norm_map = {}
    for part in freq.index:
        if part in norm_map:
            continue
        if part.endswith('S'):
            singular = part[:-1]
            if singular in freq and singular not in norm_map:
                if freq[part] >= freq[singular]:
                    norm_map[singular] = part
                else:
                    norm_map[part] = singular
    nospace_map = {}
    for part in sorted(freq.index, key=lambda x: -freq[x]):
        key = part.replace(' ', '').replace('-', '')
        if key not in nospace_map:
            nospace_map[key] = part
        else:
            canonical = nospace_map[key]
            if part not in norm_map and part != canonical:
                norm_map[part] = canonical
    for part in freq.index:
        if part in norm_map:
            continue
        alt = part.replace(' & ', ' AND ') if ' & ' in part else part.replace(' AND ', ' & ')
        if alt != part and alt in freq:
            canonical = part if freq[part] >= freq[alt] else alt
            variant = alt if canonical == part else part
            if variant not in norm_map:
                norm_map[variant] = canonical
    return norm_map

def enrich_from_remarks(df):
    def extract_maker(r):
        if not r or not isinstance(r, str): return None
        m = re.search(r'Maker\s*:\s*"?\s*([^";\n]+)', r, re.IGNORECASE)
        return m.group(1).strip().strip('"').strip() if m else None
    def extract_model(r):
        if not r or not isinstance(r, str): return None
        for pat in [r'(?:^|[,;\s(])Model\s*:\s*"?\s*([^";\n)]+)',
                    r'\(MODEL\s*:\s*([^)]+)\)',
                    r'Gear Type\s*:\s*"?\s*([^";\n]+)',
                    r'(?:^|[;\s])Type\s*:\s*"?\s*([^";\n]+)']:
            m = re.search(pat, r, re.IGNORECASE)
            if m:
                v = m.group(1).strip().strip('"').strip()
                if v: return v
        return None
    if 'Remarks' not in df.columns:
        return df
    has_remark = df['Remarks'].notna() & (df['Remarks'].astype(str).str.strip() != '')
    empty_maker = df['Maker'].isna() | (df['Maker'].astype(str).str.strip() == '')
    empty_model = df['Model / Type'].isna() | (df['Model / Type'].astype(str).str.strip() == '')
    if (has_remark & empty_maker).any():
        extracted = df.loc[has_remark & empty_maker, 'Remarks'].apply(extract_maker)
        df.loc[has_remark & empty_maker & extracted.notna(), 'Maker'] = extracted[extracted.notna()]
    if (has_remark & empty_model).any():
        extracted = df.loc[has_remark & empty_model, 'Remarks'].apply(extract_model)
        df.loc[has_remark & empty_model & extracted.notna(), 'Model / Type'] = extracted[extracted.notna()]
    return df

# ── Excel 格式輸出 ───────────────────────────────────────────────
def write_excel(df, path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'lube_chart'
    ws.freeze_panes = 'A2'

    all_cols = COLS + ['Source', 'source_file']
    # 標題列
    for ci, col in enumerate(all_cols, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        c.fill = PatternFill('solid', fgColor=HDR_BG)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for ri, (_, row) in enumerate(df.iterrows(), start=2):
        bg = ODD_BG if ri % 2 == 1 else EVEN_BG
        for ci, col in enumerate(all_cols, 1):
            val = str(row[col]) if pd.notna(row.get(col, '')) else ''
            c = ws.cell(row=ri, column=ci, value=val)
            # Source 欄顏色
            if col == 'Source':
                color = LC_COLOR
                c.font = Font(name='Arial', size=9, color=color, bold=True)
            else:
                c.font = Font(name='Arial', size=9)
            c.fill = PatternFill('solid', fgColor=bg)
            c.alignment = Alignment(vertical='center', wrap_text=True)

    # 欄寬自動調整
    for ci, col in enumerate(all_cols, 1):
        if col in df.columns:
            max_len = max((len(str(v)) for v in df[col].fillna('') if v), default=len(col))
        else:
            max_len = len(col)
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len + 2, 10), 50)

    wb.save(path)

# ── 正規化合併報告 ─────────────────────────────────────────────
def _write_normalize_report(maker_groups, model_groups):
    """產出 output/normalize_report.xlsx，列出 Maker / Model 被合併的群組。"""
    out_path = os.path.join(OUT_DIR, 'normalize_report.xlsx')
    os.makedirs(OUT_DIR, exist_ok=True)
    wb = Workbook()

    def fill_sheet(ws, title, groups):
        ws.title = title
        headers = ['key', 'canonical', 'variant_count', 'total_rows', 'variants']
        ws.append(headers)
        for g in groups:
            ws.append([
                g['key'],
                g['canonical'],
                len(g['variants']),
                g['total_rows'],
                ' | '.join(g['variants']),
            ])
        # 標題列樣式
        hdr_font = Font(color='FFFFFF', bold=True)
        hdr_fill = PatternFill('solid', fgColor=HDR_BG)
        for c in ws[1]:
            c.font = hdr_font
            c.fill = hdr_fill
        ws.freeze_panes = 'A2'
        widths = [16, 28, 14, 12, 80]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    fill_sheet(wb.active, 'maker_merged', maker_groups)
    fill_sheet(wb.create_sheet(), 'model_merged', model_groups)
    wb.save(out_path)
    print(f"  正規化報告：{out_path}（Maker {len(maker_groups)} 群組、Model {len(model_groups)} 群組）")


# ── Log 工具 ────────────────────────────────────────────────────
def append_log(summary):
    os.makedirs(LOG_DIR, exist_ok=True)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(summary + '\n')

# ── 主流程 ──────────────────────────────────────────────────────
def main():
    print("=" * 50)
    print("▶  process_lube_chart.py 開始執行")
    print("=" * 50)

    reg   = load_registry(REG_FILE)
    freg  = load_failed_registry()
    now   = datetime.now()

    # 掃描所有 CSV
    all_csvs = sorted(glob.glob(os.path.join(LC_DIR, '*.csv')))
    if not all_csvs:
        print(f"⚠️  找不到 CSV 檔案：{LC_DIR}")
        return

    # 分類：需處理 vs 跳過
    to_process = []
    skipped    = []
    for fpath in all_csvs:
        fname = os.path.basename(fpath)
        needed, reason = needs_processing(fname, fpath, reg)
        if needed:
            to_process.append((fname, fpath, reason))
        else:
            skipped.append(fname)

    # 加入 retryable 失敗檔
    retryable = [
        k for k, v in freg.get('failed_files', {}).items()
        if v.get('source') == 'lube_chart' and
           v.get('status') == 'retryable' and
           v.get('retry_count', 0) < 3
    ]
    for key in retryable:
        fname = os.path.basename(key)
        fpath = os.path.join(LC_DIR, fname)
        if os.path.exists(fpath) and fname not in [x[0] for x in to_process]:
            to_process.append((fname, fpath, 'retry'))

    print(f"CSV 總數：{len(all_csvs)}  ｜  待處理：{len(to_process)}  ｜  跳過：{len(skipped)}")

    if not to_process:
        print("✓ 無需更新，所有檔案未變更。")
        return

    # 讀取現有 master
    if os.path.exists(MASTER):
        try:
            df_master = pd.read_excel(MASTER, sheet_name='lube_chart')
            print(f"✓ 載入現有 master：{len(df_master)} 列")
        except Exception:
            df_master = pd.DataFrame(columns=COLS + ['Source', 'source_file'])
    else:
        df_master = pd.DataFrame(columns=COLS + ['Source', 'source_file'])

    # 處理統計
    stats = {'added': [], 'updated': [], 'failed': []}
    read_cols = COLS + ['Remarks']

    for fname, fpath, reason in to_process:
        try:
            print(f"\n  處理：{fname}  [{reason}]")
            df = pd.read_csv(fpath, sep=';', encoding='utf-8-sig', low_memory=False,
                             usecols=lambda c: c in read_cols)
            if 'Remarks' not in df.columns:
                df['Remarks'] = ''
            df = enrich_from_remarks(df)
            df = df[COLS].copy()

            # 全大寫 + strip
            for col in COLS:
                df[col] = df[col].fillna('').astype(str).str.strip().str.upper()

            df['source_file'] = fname
            df['Source']      = 'LUBE CHART'

            # 從 master 移除舊資料（若已存在）
            if fname in df_master.get('source_file', pd.Series(dtype=str)).values:
                df_master = df_master[df_master['source_file'] != fname]

            df_master = pd.concat([df_master, df], ignore_index=True)

            # 更新 registry
            stat = os.stat(fpath)
            reg['processed_files'][fname] = {
                'processed_at': now.isoformat(timespec='seconds'),
                'file_size_bytes': stat.st_size,
                'file_mtime': datetime.fromtimestamp(stat.st_mtime).isoformat(timespec='seconds'),
                'sha256': sha256(fpath)
            }
            # 從失敗 registry 移除（若曾失敗）
            fail_key = f'lube_chart/{fname}'
            if fail_key in freg.get('failed_files', {}):
                del freg['failed_files'][fail_key]

            if reason == 'retry':
                stats['added'].append(f'{fname} ✓（retry 成功）')
            elif reason == 'new':
                stats['added'].append(f'{fname} ✓')
            else:
                stats['updated'].append(f'{fname} ✓（{reason}）')

        except Exception as e:
            error_msg = str(e)
            print(f"  ❌ 失敗：{error_msg}")
            fail_key = f'lube_chart/{fname}'
            if fail_key not in freg.setdefault('failed_files', {}):
                freg['failed_files'][fail_key] = {
                    'source': 'lube_chart', 'first_failed_at': now.isoformat(timespec='seconds'),
                    'last_failed_at': now.isoformat(timespec='seconds'),
                    'retry_count': 1, 'error_type': 'FILE_READ_ERROR',
                    'error_message': error_msg, 'status': 'retryable'
                }
            else:
                freg['failed_files'][fail_key]['retry_count'] += 1
                freg['failed_files'][fail_key]['last_failed_at'] = now.isoformat(timespec='seconds')
                if freg['failed_files'][fail_key]['retry_count'] >= 3:
                    freg['failed_files'][fail_key]['status'] = 'manual_required'
            stats['failed'].append(fname)
            continue

    # 全域標準化（Maker + Part）
    print("\n  套用全域標準化規則...")
    maker_norm = build_maker_norm_map(df_master['Maker'])
    df_master['Maker'] = df_master['Maker'].map(lambda x: maker_norm.get(x, x))
    print(f"  Maker 標準化：{len(maker_norm)} variants")

    part_norm = build_part_norm_map(df_master['Part to be lubricated'])
    df_master['Part to be lubricated'] = df_master['Part to be lubricated'].map(lambda x: part_norm.get(x, x))
    print(f"  Part 標準化：{len(part_norm)} variants")

    # 移除 Model 末端數量描述（(3 SETS)、(X3)、(2)、 300 EA 等）
    before_models = df_master['Model / Type'].copy()
    df_master['Model / Type'] = df_master['Model / Type'].map(strip_quantity_descriptor)
    qty_stripped = (before_models != df_master['Model / Type']).sum()
    print(f"  Model 數量字尾移除：{qty_stripped} 列")

    # 過濾無效型號
    before = len(df_master)
    df_master = df_master[~df_master['Model / Type'].apply(is_invalid_model)]
    print(f"  過濾無效型號：{before - len(df_master)} 列移除")

    # 排除 TALUSIA LS 25
    before = len(df_master)
    df_master = df_master[~df_master['Lubricant'].str.contains('TALUSIA LS 25', na=False)]
    print(f"  排除 TALUSIA LS 25：{before - len(df_master)} 列移除")

    # Maker / Model 正規化（方案 A：規則式比對鍵 + 最高頻原文為標準形）
    new_makers, maker_groups = canonicalize_column(df_master['Maker'], maker_key)
    new_models, model_groups = canonicalize_column(df_master['Model / Type'], model_key)
    df_master['Maker'] = new_makers
    df_master['Model / Type'] = new_models
    print(f"  Maker 正規化：合併 {len(maker_groups)} 群組")
    print(f"  Model 正規化：合併 {len(model_groups)} 群組")

    # 寫出正規化合併報告
    try:
        _write_normalize_report(maker_groups, model_groups)
    except Exception as e:
        print(f"  ⚠️  正規化報告寫入失敗（不影響主流程）：{e}")

    # 去重
    before = len(df_master)
    df_master = df_master.drop_duplicates(subset=DEDUP_KEYS)
    print(f"  去重：{before - len(df_master)} 列移除，剩餘 {len(df_master)} 列")

    # 排序
    df_master = df_master.sort_values('source_file').reset_index(drop=True)

    # 寫入 master Excel（WRITE_ERROR 中止）
    try:
        os.makedirs(OUT_DIR, exist_ok=True)
        write_excel(df_master, MASTER)
        print(f"\n✓ lube_chart_master.xlsx 寫入完成：{len(df_master)} 列")
    except Exception as e:
        print(f"\n❌ WRITE_ERROR：{e}")
        print("⚠️  中止執行，避免 master 資料不一致。")
        return

    # 儲存 registry
    save_registry(reg, REG_FILE)
    save_failed_registry(freg)

    # manual_required 警告
    manual_req = [k for k, v in freg.get('failed_files', {}).items()
                  if v.get('source') == 'lube_chart' and v.get('status') == 'manual_required']

    # 寫 log
    log_lines = [
        '=' * 40,
        f'時間：{now.strftime("%Y-%m-%d %H:%M:%S")}',
        '動作：更新Lube Chart',
        '-' * 40,
    ]
    if stats['added']:   log_lines.append(f'新增處理：{", ".join(stats["added"])}')
    if stats['updated']: log_lines.append(f'重新處理：{", ".join(stats["updated"])}')
    if skipped:          log_lines.append(f'跳過：{len(skipped)} 個檔案（未變更）')
    if stats['failed']:
        log_lines.append(f'失敗（本次）：')
        for f in stats['failed']:
            log_lines.append(f'  - {f}')
    if manual_req:
        log_lines.append(f'\n⚠️  需人工介入（manual_required）：')
        for k in manual_req:
            log_lines.append(f'  - {k}')
    log_lines.append(f'\n輸出：output/lube_chart_master.xlsx（共 {len(df_master)} 列）')
    log_lines.append('=' * 40)
    append_log('\n'.join(log_lines))

    # 摘要輸出
    print('\n' + '=' * 50)
    print('📊 執行摘要')
    print('-' * 50)
    if stats['added']:   print(f'新增：{len(stats["added"])} 個檔案')
    if stats['updated']: print(f'更新：{len(stats["updated"])} 個檔案')
    if skipped:          print(f'跳過：{len(skipped)} 個檔案')
    if stats['failed']:  print(f'失敗：{len(stats["failed"])} 個檔案')
    print(f'輸出：lube_chart_master.xlsx（{len(df_master):,} 列）')
    if manual_req:
        print('\n⚠️  以下檔案需要人工介入（已失敗 3 次）：')
        for k in manual_req:
            print(f'   - {k}')
    print('=' * 50)

if __name__ == '__main__':
    main()
