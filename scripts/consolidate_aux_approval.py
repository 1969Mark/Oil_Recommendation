"""Consolidate OEM_data/Auxiliary approval_.xlsx into a single sheet.

Mapping:
  EQUIPMENT             = sheet name
  MAKER                 = Manufacturer
  TYPE                  = Pattern
  Part to be lubricated = Operating Conditions
  Lubricant             = first non-empty among Mineral / Synthetic / Grease
Filter: Present on listed Maker's List == "Y"
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SRC = Path("OEM_data/Auxiliary approval_.xlsx")
OUT = Path("output/auxiliary_approval_consolidated.xlsx")

OUT_HEADERS = ["EQUIPMENT", "MAKER", "TYPE", "Part to be lubricated", "Lubricant"]


def norm(v):
    if v is None:
        return ""
    return str(v).strip()


def find_header_row(rows):
    for i, r in enumerate(rows[:5]):
        if r and any(c and "Manufacturer" in str(c) for c in r):
            return i
    return None


def resolve_columns(main_row, sub_row):
    """Return dict mapping logical names to column index."""
    cols = {}
    for idx, val in enumerate(main_row):
        v = norm(val).lower()
        if v == "manufacturer":
            cols["maker"] = idx
        elif v == "pattern":
            cols["type"] = idx
        elif v.startswith("operating condition"):
            cols["part"] = idx
        elif v.startswith("present on listed"):
            cols["filter"] = idx
    for idx, val in enumerate(sub_row):
        v = norm(val).lower()
        if v.startswith("mineral"):
            cols["mineral"] = idx
        elif v.startswith("synthetic"):
            cols["synthetic"] = idx
        elif v == "grease":
            cols["grease"] = idx
    return cols


def pick_lubricant(row, cols):
    for key in ("mineral", "synthetic", "grease"):
        if key in cols:
            v = norm(row[cols[key]])
            if v:
                return v
    return ""


def main():
    wb_src = openpyxl.load_workbook(SRC, data_only=True)
    out_rows = []
    skipped_sheets = []

    for sheet_name in wb_src.sheetnames:
        ws = wb_src[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        hdr_idx = find_header_row(rows)
        if hdr_idx is None:
            skipped_sheets.append((sheet_name, "no header"))
            continue
        main_row = rows[hdr_idx]
        sub_row = rows[hdr_idx + 1] if hdr_idx + 1 < len(rows) else ()
        cols = resolve_columns(main_row, sub_row)
        if "filter" not in cols or "maker" not in cols:
            skipped_sheets.append((sheet_name, f"missing cols: {cols}"))
            continue

        data_start = hdr_idx + 2
        kept = 0
        for r in rows[data_start:]:
            if not r or all(c is None for c in r):
                continue
            if norm(r[cols["filter"]]).upper() != "Y":
                continue
            maker = norm(r[cols["maker"]])
            type_ = norm(r[cols.get("type", -1)]) if "type" in cols else ""
            part = norm(r[cols.get("part", -1)]) if "part" in cols else ""
            lub = pick_lubricant(r, cols)
            if not (maker or type_ or part or lub):
                continue
            out_rows.append([sheet_name, maker, type_, part, lub])
            kept += 1
        print(f"{sheet_name}: kept {kept}")

    if skipped_sheets:
        print("Skipped:", skipped_sheets)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "consolidated"

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)
    even_fill = PatternFill("solid", fgColor="EBF3FB")

    ws_out.append(OUT_HEADERS)
    for c in ws_out[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for i, row in enumerate(out_rows, start=2):
        ws_out.append(row)
        if i % 2 == 0:
            for c in ws_out[i]:
                c.fill = even_fill

    for col_idx, h in enumerate(OUT_HEADERS, start=1):
        max_len = max([len(h)] + [len(str(r[col_idx - 1])) for r in out_rows]) + 2
        ws_out.column_dimensions[ws_out.cell(row=1, column=col_idx).column_letter].width = min(max(max_len, 10), 50)
    ws_out.freeze_panes = "A2"

    wb_out.save(OUT)
    print(f"\nWrote {len(out_rows)} rows -> {OUT}")


if __name__ == "__main__":
    main()
